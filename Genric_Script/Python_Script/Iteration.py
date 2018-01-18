#!/usr/bin/env python

import AutoItLibrary     #this library is imported to automate the windows GUI
import win32com.client     #this library is imported to automate the windows GUI
import subprocess
import sys,os
import shutil   
import paramiko   
import openpyxl
import ConfigParser
Config = ConfigParser.ConfigParser()
from xml.dom import minidom
from openpyxl import load_workbook
reload(sys)  
sys.setdefaultencoding('UTF-8')
autoit=win32com.client.Dispatch("AutoItX3.Control")  # this is used to move file on windows
#import pcapparser   #this is imported for reading pcap file 
#import paramiko    #this is imported for connecting to remote server and copy files to local machine
#from subprocess import Popen, PIPE

autoit=win32com.client.Dispatch("AutoItX3.Control")

 
				
def rev_tab(num):
    
    for i in range(0,int(num)):	 
        autoit.Send("+{TAB}")
        if i==num: 
            break

#The below func is used to rename the file
def rename_file(path_of_file,src,dst):
		os.chdir(path_of_file)
		os.rename(src,dst)
		print "the dir is: %s" %os.listdir(os.getcwd())
		
def get_string():
	
		o = open('C:\\Users\\Techm\\Desktop\\test.txt','r+')
		r= o.read()
		s=r.split()
		return s[12]
	
def join_string():
		o=open('C:\\Users\\Techm\\Desktop\\ip.txt','r+')
		r= o.read()
		seq=(r,"service network 1")
		s = " ".join(seq)
		#print s
		return s
		
def RQM_status1():
		o=open('C:\\Users\\Techm\\Sukesh-Automation\\RobotFramework\\Volte\\VOLTE_TestCases\\Results\\output.xml','r+')
		r= o.readlines()
	#line_num=0
	#num=[]
	#inde=[]
		status="FAIL"
	
		for line in r:
			#line_num+=1
			a = line.split()
			i= line.find('fail="0"')
			#inde.append(i)
			#for i in inde:
			if i != -1 :  
				status = "PASS"
			#print status
		
		return status

def	tup_to_list(t):	
		l=list(t)
		x= l[0]
		y= l[1]
		return x,y
	
def RQM_status2(status):
	
		o=open('C:\\Users\\Techm\\Sukesh-Automation\\RobotFramework\\RQM-Integration\\status.txt','r+')
		o.write(status)		
		
		#if inde >=0 : print "PASS"
		#else : print "FAIL"
		#print a
		#if line.find('fail="0"')>= 0: 
			#num.append(line_num)			
			#i=line.find('fail="0"')
			#print i
			
			
			
		#print num,inde
		
	#return inde
		#b=a.index("fail=")
		#print list
			
	
	#if index == -1: 
		#print 'P'
		#return 'Pass'
	#else :
		#print 'F'
		#return 'Fail'
	
		
#def	Pcap_Parse():

	
	
	
	
	
		
#def  copy_file():		
#client = paramiko.SSHClient()
#client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
#client.connect('172.18.0.237', username='root', password='!bootstrap')
#ssh_stdin, ssh_stdout, ssh_stderr = client.exec_command('ls')
#print "output", ssh_stdout.read()
#proc = Popen(['sftp','root@172.18.0.237', 'stop'], stdin=PIPE)
#proc.stdin.write('!bootstrap\n')
#proc.stdin.flush()
#sftp = client.open_sftp()
#rf = sftp_client.open('remote_filename')
#rf.readline()
#print rf
#return rf
#print sftp.get_channel()
#print sftp.getcwd()
#sftp.get("ims1_trace_pcscf.pcap")
#sftp.close()
#ssh.close()

#if __name__ == '__main__':

	#import sys
	#import Script from Iteration
	
	
def Imagereport():
	os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Imageupload_Report.bat")	

def Onboardingreport():
	os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Onboarding_Report.bat")	

def Testingreport():
		
		path= "C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/TestingResults"
		os.chdir(path)
		os.rename('report.html','IMSreport.html')
		os.rename('output.xml','IMSoutput.xml')
		os.rename('log.html','IMSlog.html')
		#print "the dir is: %s" %os.listdir(os.getcwd())	
		os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Testing_Report.bat")	
def Get_data( filename):
	
	client = paramiko.SSHClient()
	client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	client.connect('10.53.214.10', username='root', password='techm123')
	sftp = client.open_sftp()
	rf = sftp.open(filename)
	line= rf.read()
	l=line.strip()
	#print l
	return l	
def get_string():
		o = open('C:\\Users\\Techm\\Downloads\\pwd.txt','r+')
		r= o.read()
		s= r.split()
		print s
		a=s[271]
		return a	
def join_string():
		o=open('C:\\Users\\Admin\\Desktop\\Sukesh-Automation\\RobotFramework\\perimeta\\virt-serv1_portip.txt','r+')
		r= o.read()
		s= r.split()
		seq=(s[0],"service-network 1")
		s = " ".join(seq)
		#print s
		return s
	
def join_string1():
		o=open('C:\\Users\\Admin\\Desktop\\Sukesh-Automation\\RobotFramework\\perimeta\\virt-serv2_portip.txt','r+')
		r= o.read()
		s= r.split()
		seq=(s[0],"service-network 2")
		s = " ".join(seq)
		#print s
		return s
def join_string2(id,ip):
		id1=id.strip()
		list=[]
		list.append(id1)
		list.append(ip)
		seq1=(list[0],"--allowed-address-pairs type=dict list=true ip_address=")
		a= " ".join(seq1)
		seq2=(a,list[1]) 
		s= "".join(seq2)
		return s		
def join_string3(id,ip1,ip2,ip3):
		id1=id.strip()
		list=[]
		list.append(id1)
		IP1=ip1.strip()
		IP2=ip2.strip()
		IP3=ip3.strip()
		list.append(IP1)
		list.append(IP2)
		list.append(IP3)
		#print list
		seq1=(list[0],"--allowed-address-pairs type=dict list=true")
		a= " ".join(seq1)
		seq2=(a,'ip_address=')
		s2= " ".join(seq2)
		seq3=(s2,list[1])
		s3="".join(seq3)
		seq4=(s3,'ip_address=')
		s4= " ".join(seq4)
		seq5=(s4,list[2])
		s5="".join(seq5)
		seq6=(s5,'ip_address=')
		s6= " ".join(seq6)
		seq7=(s6,list[3])
		s7="".join(seq7)
		print s7
		return s7
def join_string4(ip):
		seq1="# signaling-local-address ipv4"
		seq2=(seq1,ip)
		s= " ".join(seq2)
		return s		
def join_string5(id):
	i=id.strip()
	seq1=(i,"--allowed-address-pairs type=dict list=true ip_address=0.0.0.0/1 ip_address=128.0.0.0/1 ip_address=::/1 ip_address=8000::/1")
	a= " ".join(seq1)
	#b= (id,seq1)
	return a	
#def tabloop(num):    
	for i in range(0,int(num)):	 
		autoit.Send("{TAB}")
		if i==num: 
			break
def rename_file(path_of_file,src,dst):	
	os.chdir(path_of_file)
	os.rename(src,dst)
	print "the dir is: %s" %os.listdir(os.getcwd())

def Imagereport():
	os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Imageupload_Report.bat")	

def Onboardingreport():
	os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Onboarding_Report.bat")	

def Testingreport():
		#os.chdir("C:\\Users\\Admin\\Desktop\\Sukesh-Automation\\RobotFramework\\perimeta\\TestingResults")
		#os.rename("C:\Users\Admin\Desktop\Sukesh-Automation\RobotFramework\perimeta\batfiles\report.html","IMSreport")
		#print "the dir is: %s" %os.listdir(os.getcwd())	
		os.system("C:/Users/Admin/Desktop/Sukesh-Automation/RobotFramework/perimeta/batfiles/Testing_Report.bat")		
		
def parseExcelData(path,filename,ServerName,requiredvalue):
    required_value = ""   
    data_file = os.path.join(path, filename)
    wb = load_workbook(filename=data_file)
    sheets = ["TestData"]
    for sheet in sheets:
        ws = wb.get_sheet_by_name(sheet)
        for row in ws.iter_rows('A1:A6'):
            for cell in row:					
                if (cell.value == "ServerName" and ServerName == "Server1" ):
                    if (requiredvalue == "ServerIP"):
                        required_value= ws['B2'].value
                        #print ws['B2'].value
                    if (requiredvalue == "UserName"):
                        required_value= ws['C2'].value
                    if (requiredvalue == "Password"):
                        required_value= ws['D2'].value
                        
                if (cell.value == "ServerName" and ServerName == "Server2" ):
                    if (requiredvalue == "ServerIP"):
                        required_value= ws['B3'].value
                    if (requiredvalue == "UserName"):
                        required_value= ws['C3'].value
                    if (requiredvalue == "Password"):
                        required_value= ws['D3'].value
                return required_value;
				
				
def parsexmldata():

	xmldoc = minidom.parse('C:\\Users\\Admin\\Desktop\\testfile.xml')
	itemlist = xmldoc.getElementsByTagName('item')
	for s in itemlist:
		print "Server name : ", s.attributes['server'].value
		return s.attributes['server'].value
		
#def parseconfdata(section):
	Config.read("C:\\Users\\Admin\\Desktop\\testfile.conf")
	Config.sections()
	dict1={}
	options = Config.options(section)
	for option in options:
		try:
			dict1[option] = Config.get(section, option)
			strng = str(dict1[option])
			if dict1[option] == -1:
				DebugPrint("skip: %s" % option)
		except:
			print("exception on %s!" % option)
			dict1[option] = None
		return strng	